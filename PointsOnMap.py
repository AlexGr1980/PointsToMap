# Вывод на карту координат из файла *.xlsx
# 1 столбец - Широта, 2 столбец - Долгота, 3 столбец - Метка
# 1 столбец - Ш=44°18'22,81''N, 2 столбец - Д=38°35'27,23''E, 3 столбец - Метка, если пусто, то это текущая координата поля, если заполнено, то это метка и поле закрывается

import os
import re
import math
from loguru import logger 
import openpyxl
import folium
from folium.features import DivIcon

# Создаем класс "Координата" - Широта и долгота точки
class clCoordinate:
    def __init__(self, Latitude, Longitude):
        self.Latitude = Latitude
        self.Longitude = Longitude

    def display_info(self):  
        print(f"Широта: {self.Latitude}, Долгота: {self.Longitude}") 

# Функция, которая анализирует состав поля координат и преобразует их в координаты формата: 48,56456456
def TransformCoordinate(value)->float:
    logger.info("---Старт пересчета координат---")
    transCoord = 0.0
    pattern = re.compile("=(\d*)°")
    gradus = pattern.findall(value)
    transCoord +=float(gradus[0])
    logger.info(f"Градус: {gradus[0]}. Пересчет: {transCoord}")

    pattern = re.compile("°(\d*)'")
    minute = pattern.findall(value)
    transCoord +=float(minute[0])/60.0
    logger.info(f"Минута: {minute[0]}. Пересчет: {transCoord}" )
        
    pattern = re.compile("'(.*)'{2}")
    second_buff = pattern.findall(value)
    second = re.sub(",", ".", second_buff[0])
    transCoord +=float(second)/3600.0
    logger.info(f"Секунда: {second}. Пересчет: {float(second[0])/3600.0}" )
    logger.info("---Конец пересчета координат---")
    logger.info(f"================================")

    return transCoord

# Функция, которая выводит координаты на карту
def SetInfoToMap(Fields):
    area =None
    logger.info(f"===============================")
    logger.info(f"---Старт работы с картой---")
    area = folium.Map(location=[Fields[0]['1'].Latitude, Fields[0]['1'].Longitude], width='100%' , height = '100%', titles = "OpenStreetMap" ,zoom_start=10)

    count = 1
    for Element in Fields:
        points = []
        for key, value in Element.items():
            if key != "Label":
               
                logger.debug(f"Метка: {Element['Label']}. Координата {key}: широта: {value.Latitude}, долгота: {value.Longitude}")
            
                points.append([value.Latitude, value.Longitude])
                folium.CircleMarker(location=[value.Latitude, value.Longitude], radius = 3, popup=f"Label", fill_color="red", color="gray", fill_opacity = 1).add_to(area)
            else:
                #Замыкаем координаты
                points.append(points[0])
                #Рисование название "Поля"
                folium.map.Marker(location=[Element['1'].Latitude, Element['1'].Longitude], icon= DivIcon(icon_size=(250,50), icon_anchor=(10,0),html=f'<div style="font-size: 10pt">"{value}"</div>',)).add_to(area)
    
        #Рисование полилиний
        folium.PolyLine(points).add_to(area)    

    
    #Координаты газопровода отдельные Джубга-Лахаревское-Сочи
    #loc_gas=[(44.32,38.6551),(44.3084,38.6631),(44.2667,38.7819),(44.1845,38.8585),(44.1251,39.0063),(43.6016,39.6303),(43.5867,39.7125)]   
    #folium.PolyLine(loc_gas, color='red', weight=1, opacity=0.8).add_to(area)

    area.save(r"Poins_to_Map.html")

    logger.info(f"===============================")
    logger.info(f"---Конец работы с картой---")

def ReadCoordinates(FileName, NumericList) -> dict:

    logger.info(f"---Старт считывания координат из файла---")
    logger.info(f"===============================")
    WorkBook = openpyxl.load_workbook(FileName)
    Sheets = WorkBook.sheetnames
    WorkSheet = WorkBook[Sheets[NumericList]] # Активация 1 листа, где хранятся координаты
    MaxRows = WorkSheet.max_row
    MaxColumns = WorkSheet.max_column

    Fields = [] #[{label:"1", 1:[,], 2:[,], 3:[,]}, {label:"2", 1:[,], 2:[,], 3:[,]}]
    Field = {}
    
    Count_coord =1
    for i in range(1, MaxRows+1):
        for j in range(1, MaxColumns+1):  
            match j:
                case 1:
                    Latitude = TransformCoordinate(WorkSheet.cell(row=i, column=j).value)
                case 2:
                    Longitude = TransformCoordinate(WorkSheet.cell(row=i, column=j).value)
                    Coordinate = clCoordinate(Latitude, Longitude) #сохранили координату
                    Field[str(Count_coord)] = Coordinate
                case 3:
                    Label = WorkSheet.cell(row=i, column=j).value
                    if Label != None:
                        Field["Label"] = Label
                        # Проверка наполнения словаря
                        for key in Field.keys():
                            if key !="Label":
                                logger.debug(f"Элемент: {key}, широта: {Field[key].Latitude}")
                        Fields.append(Field)

                        # Обнуляем проход
                        Count_coord =0
                        Latitude = None
                        Longitude = None
                        Label = None
                        Field = {}

                    Count_coord +=1 # Счетчик координат отдельных участков
    logger.info(f"===============================")
    logger.info(f"---Конец считывания координат из файла---")
    return Fields

@logger.catch # Отлавлливает ошибки в коде
def main():
    LogFile = "LogFile.log"
    os.system("cls") # Очистка консоли

    # Удаляем старый log- файл
    if os.access(LogFile, os.F_OK) == True:
        os.remove(LogFile)

    logger.remove(0)
    logger.add(LogFile, format="| {time:DD.MM.YY hh:mm:ss} | {message} |", level="DEBUG", rotation = "1Mb", compression="zip")
    logger.info("Старт программы")
    logger.info(f"================================")
    print("---Старт программы---")


    FileName = "Координаты.xlsx"
    NumericList = 0 #Номер листа на котором размещены координаты. Д.б. на 1 листе.
    Fields = []
    Fields = ReadCoordinates(FileName, NumericList)
    SetInfoToMap(Fields)

    logger.info(f"================================")
    logger.info("Программа выполнена")
    print("---Программа выполненны---")

if __name__ == "__main__":
     main()